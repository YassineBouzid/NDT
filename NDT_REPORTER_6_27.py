# -*- coding: utf-8 -*-
# encoding=utf8
import sys
reload(sys)
sys.setdefaultencoding('utf8')

import tkinter as tk
from tkinter import ttk,messagebox
import os
import sys
import time
from datetime import date
from datetime import datetime,timedelta

from tkinter import *
import tkinter.font as font
import base64
os.environ['CUDA_VISIBLE_DEVICES'] = '0'
from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import win32api,win32con
from tkinter.scrolledtext import ScrolledText
import openpyxl
from openpyxl.drawing.image import Image as xlimg
import sqlite3 as sq
import glob
import shutil

import qrcode
from PIL import Image

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.getcwd()
    return os.path.join(base_path, relative_path)


j=int()
j=1
k=int()
k=1
max_line = 25
obser_var =0
x=10
insertion_var=True
observation=''
observation_list=[]

time_value=time.strftime("%d-%m-%y")
print(time_value)



timing = datetime.now()
if 21<=timing.hour<24 :
    time_value=datetime.now()+timedelta(1)
else:
    time_value=datetime.now()
time_value=time_value.strftime("%d-%m-%y")
print(time_value)







####################################################################################################################################################################################################################################################################################################################################################################################################################
########################################################################---------------------PYDICOME READER-------------------------#################################################################################################################################################################################################################################################################################################################


# TO DO: FINISH STORING IN DATABSE


def convertToBinaryData(filename):
    # Convert digital data to binary format
    with open(filename, 'rb') as file:
        blobData = file.read()
    return blobData

def create_atable(NAME_DATABASE,NAME_TABLE,REPORT,col1,col2,col3,col4):
    cnn =sq.connect('{}.db'.format(NAME_DATABASE))
    c = cnn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS {} ({} text, {} text, {}  integer,{}  integer, {}  text)".format(NAME_TABLE,REPORT,col1,col2,col3,col4))
    cnn.commit()
    cnn.close()
    print("done!")

def insert_one_record(NAME_DATABASE,NAME_TABLE,NAME_REPORT,REC1,REC2,REC3,REC4):
    cnn =sq.connect('{}.db'.format(NAME_DATABASE))
    c = cnn.cursor()
    
    c.execute("INSERT INTO {} VALUES (?,?,?,?,?)".format(NAME_TABLE),(NAME_REPORT,REC1,REC2,REC3,REC4))
    cnn.commit()
    cnn.close()
    print("done!")
    
def select_by_pipe_name(NAME_DATABASE,NAME_TABLE,NAME_REPORT, PIPENAME):# done
    cnn =sq.connect('{}.db'.format(NAME_DATABASE))
    c = cnn.cursor()
    c.execute("SELECT  rowid, * FROM {} WHERE TUBE = '{}'and REPORT like '%{}%'".format(NAME_TABLE,PIPENAME,PROJECT))
    rows = c.fetchall()
    for row in rows:
        
        print(row[:3])
        print("")
    return rows

def update_arow(NAME_DATABASE,NAME_TABLE,rowid,NAME_REPORT,REC1,REC2,REC3,REC4):
    cnn =sq.connect('{}.db'.format(NAME_DATABASE))
    c = cnn.cursor()
    c.execute("UPDATE {} SET  REPORT = (?), TUBE = (?), KRASHAGE = (?), KRASHAGE_OPR = (?), OBSERVATION = (?) WHERE rowid  =  (?)".format(NAME_TABLE),(NAME_REPORT,REC1,REC2,REC3,REC4,rowid))
    cnn.commit()
    cnn.close()
    print("done!")
    

def rq_genrator(string):
    global CURRENT_INSPECTEUR
    img = qrcode.make(string)
    img.save("QR.jpg")
    img = Image.open("QR.jpg")
    img = img.convert("RGBA")

    pixdata = img.load()

    width, height = img.size
    for y in range(height):
        for x in range(width):
            if pixdata[x, y] == (255, 255, 255, 255):
                pixdata[x, y] = (255, 255, 255, 0)
    new_width  = 130
    new_height = 130
    img = img.resize((new_width, new_height), Image.ANTIALIAS)
    img.save("img2.jpg", "PNG")
    
    wb = load_workbook('QR_template.xlsx')
    wb.save("QR_print.xlsx")
        
    wb = load_workbook("QR_print.xlsx")   
    ws = wb.active
    try:
        rq_image =xlimg("img2.jpg")
        ws['b6'] = "Krash: {} K".format(int(krash.get()))
        ws['a12'] = "INSP UT: {} ".format(CURRENT_INSPECTEUR)
        ws.add_image(rq_image,'b8')
        wb.save("QR_print.xlsx")
        
    except Exception as e:
        print("eeeeeeeee",e)
    win32api.ShellExecute(
        0,
        "print",
        ("QR_print.xlsx"),
        None,
        ".",
        0
        )
    return img

    


## DECLARATION OF DATABASE COLUMN NAMES:

NAME_DATABASE= "RAPPORT_UT"
NAME_TABLE = "RAPPORT_UT"
REPORT= "REPORT"
col1="TUBE"
col2="KRASHAGE"
col3="KRASHAGE_OPR"
col4="OBSERVATION"
CURRENT_INSPECTEUR=None 

####################################################################################################################################################################################################################################################################################################################################################################################################################
####################################################################################################################################################################################################################################################################################################################################################################################################################


def tube_finished(event):
    global x,j,k,ws,operators_names1,observation,path,k,POST1,time_value,getname,PROJECT,NAME_REPORT,REC1,REC2,REC3,REC4,CURRENT_INSPECTEUR
    # j== number of line
    path = str(pathE.get())
    print("path==",path)
    krashage = str(krash.get())
    getKOPdonne=""

    #START GET THE NAME AND ADD ZEROS TO HTE PIPE NAME
    global getname
    pipe_name=str(PIPE_NAME.get())

    if PIPE_NAME.get().upper()=="TEST":
        print("test is test")
        getname='TEST'
    elif len(pipe_name)==2:
        print("length====",len(pipe_name),pipe_name[1])
        #A0001
        getname = (pipe_name[0]+"0"+"0"+"0"+pipe_name[1]).upper()
        print("getname",getname)
    elif len(PIPE_NAME.get())==3:
        #A0012
        getname = (pipe_name[0]+"0"+"0"+pipe_name[1]+pipe_name[2]).upper()
    elif len(PIPE_NAME.get())==4:
        #A0123
        getname = (pipe_name[0]+"0"+pipe_name[1]+pipe_name[2]+pipe_name[3]).upper()
    elif len(PIPE_NAME.get())==5:
        #A1234
        devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-INSPECTOR RT-II 2021",fg="black",bg=color)
        getname = str(PIPE_NAME.get()).upper()
        print("getname= ",getname)
    elif ((PIPE_NAME.get()).upper().find('-BIS') != -1):
        devlabel.config(text= "BIS TUBE!!!",fg="black",bg="yellow")
        getname = str(PIPE_NAME.get()).upper()
    
    #END GET THE NAME AND ADD ZEROS TO HTE PIPE NAME
  
    if operators_names.get()=="" or operators2_names.get()=="" or POST.get()=="" or pathE.get()=="" or PIPE_NAME.get()=="" :
        print("fill up all entries")
        devlabel.config(text= "FILL UP ALL ENTRIES!",bg="yellow" )
        return
        
    if krashage=="" :
        print("you have to start the scan or set k ")#-------------------------------------------------------------------------------------------------
        devlabel.config(text= "FILL UP HOW MUCH \n CRASHES IN THE TUBE",bg="yellow" )
        return
    
    if insertion_var:
        if os.path.isfile(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value)) and j==1:
            print("Repport exists!")
            devlabel.config(text= "REPPORT EXISTS VIRIFY NAMES\n OF THE EXISTED FILES",bg="yellow" )
            #return
    if not os.path.isfile(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value)):
        wb = load_workbook('UT_template.xlsx')
        
        
        wb.save(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value))
    wb = load_workbook(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value))
    ws = wb.active
    try:
        logo =Image("logo.jpg")  
        ws.add_image(logo,'A1')
    except Exception as e:
        print("eeeeeeeee",e)
    #ws['A1'] = "Page: {}".format(k)
    ws['H1'] = "Page: {}".format(k)
    ws['H1'].font = Font(size=18)
    ws['A3'] = "Projet: {}".format(PROJECT)
    
    #ws['E2'] = 'Rapport de contrôle\n CONTROLE UT AUTO N° "{}"'.format(k)
    # set the "Equipe" and the "Post" from the form
    #EQUIPE1 POST1 operators2_names1 operators_names1
    
    ws['E4'] ='Equipe et Post de Travail: "{}"'.format(POST1)

    ############################################ set the name of the operators#################################################################################
    namee= operators_names1
    ws['A33'] ='Nom et Prénom :\n{}\n Visa:'.format(operators_names1)
    ws['F33'] ='Nom et Prénom :                         Représentant Client:\n{}\nVisa:                                                     Visa:'.format(operators2_names1)
    ############################################ set the name of the operators#################################################################################

    getK=""
    getKOP=""
    ws['E3'] = 'DATE: {}'.format(time.strftime("%d-%m-%y"))
    ws['E3'].font = Font(size=18)
    ws['A{}'.format(j+7)] = j
    ws['B{}'.format(j+7)] = getname

    #krashage
    if getname =="TEST":
        getK = "OK"
    else:
        getK = int(krash.get())
    
    #opr
    if str(krash_OP.get())=="":
        getKOP = "--"
        getKOPdonne="__"
    else:
        getKOP= int(krash_OP.get())
        getKOPdonne= int(krash_OP.get())
        
    ws['C{}'.format(j+7)] = getK
    ws['D{}'.format(j+7)] = getKOP
    ws['E{}'.format(j+7)] = "/"
    ws['F{}'.format(j+7)] = "/"
    ws['H{}'.format(j+7)] = "--"
    
    observation=(str(observation_entrey.get("1.0", tk.END))).encode('ascii', 'ignore').decode('ascii')
    observation_list.append(covert_to_excel(observation))
    observation =""
    observation_entrey.delete('1.0',"end")

    
    ws['G8']= covert_to_excel(observation_list)
    
    



#CREATE A TABLE FOR THE REPORT
    #TABLE COLUMNS
    
##    col5="ACTION"
##    col6="ZIP_IMAGES"
##    
    create_atable(NAME_DATABASE,NAME_TABLE,REPORT,col1,col2,col3,col4)
    print("table created!")
    

    
    # RECORD DECLARATIONS
    NAME_REPORT="{}_RAPPORT_UT_N_{}_{}_{}_AND_{}_{}".format(PROJECT,k,POST1,operators_names1,operators2_names1,time_value)
    #NAME_REPORT=str(NAME_REPORT).encode('ascii', 'ignore').decode('ascii')
    NAME_REPORT=NAME_REPORT.replace(' ','_')
    NAME_REPORT=NAME_REPORT.replace('-','_')
    NAME_REPORT=NAME_REPORT.replace('"','')
    #print("name report=",NAME_REPORT)
    
    REC1=str(getname)
    REC2= str(getK)
    REC3= str(getKOPdonne)
    REC4=str(covert_to_excel(observation_list)) 
    
    #one_record=(NAME_REPORT,REC1,REC2,REC3,REC4)
    ROW = select_by_pipe_name(NAME_DATABASE,NAME_TABLE,NAME_REPORT,getname)
    
    if REC3== "__":
        opstring=""
    else:
        opstring=str(REC3)+" K_OP "
        
    if CURRENT_INSPECTEUR and QR_confirmed:
        for_qr_record= covert_to_excel("INSPECTEUR UT: "+CURRENT_INSPECTEUR+" TUBE: "+str(REC1)+" SNUP: "+ str(REC2)+" K "+opstring)
        rq_genrator(for_qr_record)
    
    print(ROW)
    if ROW:
        rowid=ROW[0][0]
        #print("messagebox this file is exists with rowid= ",rowid)
        #messagebox.showinfo("FOLDER EXISTES!","THIS FOLDER IS EXISTES {}".format([R[0] for R in ROW]))
        confirmation = messagebox.askquestion("CONFIRMATION CLOTURAGE!","FICHIER EXISTE {} FOIS IN:\n {} \nVEUILEZ VRAIMENT ECRASE LA PRECEDENT DONNEE?".format(len(ROW),[R[1:-1] for R in ROW]))
        #print('confirmation',confirmation)
        if confirmation=="yes":
            update_arow(NAME_DATABASE,NAME_TABLE,rowid,NAME_REPORT,REC1,REC2,REC3,REC4)
           #rest all widget
            PIPE_NAME.delete(0,"end")
            PIPE_NAME.focus_set()
            krash.delete(0,"end")
            krash.insert(0,"0")
            krash_OP.delete(0,"end")
            observation=""
            #krash.insert(0,"0")
            #START_btn.config(text ="START",bg=btncolor)
            j+=1
            print("!!! tube finished !!!","j======",j)
            devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2021",bg=color)
            finish_tube.config(text="TUBE N°{}".format(j),bg ="RoyalBlue1")

            
        else:
            print("not agree")
            PIPE_NAME.delete(0,"end")
            PIPE_NAME.focus_set()
            krash.delete(0,"end")
            krash.insert(0,"0")
            krash_OP.delete(0,"end")
            observation=""
            #krash.insert(0,"0")
            #START_btn.config(text ="START",bg=btncolor)
            
           
            print("!!! tube finished !!!","j======",j)
            devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2021",bg=color)
            return "break"
    else:
        insert_one_record(NAME_DATABASE,NAME_TABLE,NAME_REPORT,REC1,REC2,REC3,REC4)
        print("record inserted!")
    


        
        
        wb.save(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value))
        if j >= max_line:
            win32api.ShellExecute(
            0,
            "print",
            r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value),
            None,
            ".",
            0
            )
            #rest all widget
            #os.startfile(r"{}\POST REPORT N°{} at {}.xlsx".format(path,k,time_value), 'print')
            PIPE_NAME.delete(0,"end")
            krash.delete(0,"end")
            krash.insert(0,"0")
            krash_OP.delete(0,"end")
            
            print("!!!!!!!!rapport closed!!!!!!!!")
            report_closed_btn.config(text="R-N°{}".format(k),bg = "green2")
            finish_tube.config(text="INSÈRE",bg =btncolor)
            k+=1
            j=1
            PIPE_NAME.focus_set()
            PIPE_NAME.delete(0,"end")
            krash.delete(0,"end")
            krash.insert(0,"0")
            krash_OP.delete(0,"end")
            #krash.insert(0,"0")
            #START_btn.config(text ="START",bg=btncolor)
        #rest all widget
        PIPE_NAME.delete(0,"end")
        PIPE_NAME.focus_set()
        krash.delete(0,"end")
        krash.insert(0,"0")
        krash_OP.delete(0,"end")
        observation=""
        #krash.insert(0,"0")
        #{START_btn.config(text ="START",bg=btncolor)
        j+=1
        print("!!! tube finished !!!","j======",j)
        devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2021",bg=color)
        finish_tube.config(text="TUBE N°{}".format(j),bg ="RoyalBlue1")

    
def report_closed(event):
    global k,j,observation_list
    path = str(pathE.get())
    if operators_names.get()=="" or operators2_names.get()=="" or POST.get()=="" or pathE.get()=="":
        print("fill up all entries")
        devlabel.config(text= "FILL UP ALL ENTRIES!",bg="yellow" )
        return
    if j==1:
        if not os.path.isfile(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value)):
            wb = load_workbook('UT_template.xlsx')
            wb.save(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value))
        wb = load_workbook(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value))
        ws = wb.active
        try:
            logo =Image("logo.jpg")  
            ws.add_image(logo,'A1')
        except Exception as e:
            print("eeeeeeeee",e)
        #ws['A1'] = "Page: {}".format(k)
        ws['H1'] = "Page: {}".format(k)
        ws['H1'].font = Font(size=18)
        #ws['E2'] = 'Rapport de contrôle\n CONTROLE UT AUTO N° "{}"'.format(k)
        # set the "Equipe" and the "Post" from the form
        #EQUIPE1 POST1 operators2_names1 operators_names1
        ws['E4'] ='Equipe et Post detravail:   "{}"'.format(POST1)

        ############################################ set the name of the operators#################################################################################
        namee= operators_names1
        ws['A33'] ='Nom et Prénom :\n{}\n Visa:'.format(operators_names1)
        ws['F33'] ='Nom et Prénom :                         Représentant Client:\n{}\nVisa:                                                     Visa:'.format(operators2_names1)
        ############################################ set the name of the operators#################################################################################  
        ws['E3'] = 'DATE: {}'.format(time.strftime("%d-%m-%y"))
        observation=(str(observation_entrey.get("1.0", tk.END))).encode('ascii', 'ignore').decode('ascii')
        observation_list.append(covert_to_excel(observation))
        observation =""
        observation_entrey.delete('1.0',"end")
        ws['G8']= covert_to_excel(observation_list)
        wb.save(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value))
        print("The report is empty!")

    wb = load_workbook(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value))
    ws = wb.active
    try:
        logo =Image("logo.jpg")  
        ws.add_image(logo,'A1')
    except Exception as e:
        print("eeeeeeeee",e)

    observation=(str(observation_entrey.get("1.0", tk.END))).encode('ascii', 'ignore').decode('ascii')
    observation_list.append(covert_to_excel(observation))
    observation =""
    observation_entrey.delete('1.0',"end")
    ws['G8']= covert_to_excel(observation_list)
    wb.save(r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value))
    
# PRINT THE REPPORT:
    win32api.ShellExecute(
    0,
    "print",
    r"{}\RAPPORT CONTROLE UT AUTO N {}_{}_{}.xlsx".format(path,k,POST1,time_value),
    None,
    ".",
    0
    )
    #rest all widget
    #os.startfile(r"{}\POST REPORT N°{} at {}.xlsx".format(path,k,time_value), 'print')
    PIPE_NAME.delete(0,"end")
    krash.delete(0,"end")
    krash.insert(0,"0")
    krash_OP.delete(0,"end")
    print("!!!!!!!!rapport closed!!!!!!!!")
    report_closed_btn.config(text="R-N°{}".format(k),bg = "green2")
    finish_tube.config(text="INSÈRE",bg =btncolor)
    k+=1
    j=1
    observation_list=[]
    PIPE_NAME.focus_set()
    PIPE_NAME.delete(0,"end")
    krash.delete(0,"end")
    krash.insert(0,"0")
    krash_OP.delete(0,"end")
    #krash.insert(0,"0")
    
    devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2021",bg=color)

def check5(event):
    global j
    if var7.get()!=1:
        jr_variable_lab.grid(row = 10 , column= 0, padx = 10,sticky="e")
        kr_variable_lab.grid(row = 11 , column= 0, padx = 10,sticky="e")
        jr_variable.grid(row = 10 , column= 1, padx = 10,pady= 5,sticky="W")
        kr_variable.grid(row = 11 , column= 1, padx = 10,pady=5,sticky="W")
        j_variable_btn.grid(row =17,column=1, padx = 10,pady = 10,sticky="e", columnspan = 2)
        print("var7=", var7.get(),"j=",j)
        
        if var9.get()==1 or  var10.get()==1:
            root.geometry("435x427")
        else:
            root.geometry("287x427")
                        
    else:
        jr_variable_lab.grid_forget()
        kr_variable_lab.grid_forget()
        jr_variable.grid_forget()
        kr_variable.grid_forget()
        
        if var9.get()==1 or  var10.get()==1:
            root.geometry("435x323")
            
        else:
            root.geometry("287x323")
            
        j_variable_btn.grid_forget()
        print("var7=", var7.get(),"j=",j)




        

def show_frame(frame):
    if not (PASS_INP_02.get()=="" or PASS_INP_01.get()==""):
        if not (PASS_INP_01.get()==get_password(CURRENT_INSPECTEUR)[0][0] or PASS_INP_02.get()==get_password(CURRENT_INSPECTEUR)[0][0]):
            check_update_passord1.config(bg = "red")
            check_update_passord2.config(bg = "red")
            print("mismatch password")
            return
    if CURRENT_INSPECTEUR==None:
        messagebox.showwarning("CURRENT INSPECTOR","INSERT PASSOWRD PLEASE!")
        return "break"
    if str(operators_names.get())=="" or  str(operators2_names.get())=="" or str(POST.get())=="" or str(EQUIPE.get())=="" :
        operator1_names_lab.config(bg = "orange1")
        operator2_names_lab.config(bg = "orange1")
        POST_lab.config(bg = "orange1")
        PROJECT_lab.config(bg = "orange1")
        messagebox.showwarning("ENTRIES ISSUE","FILL UP ALL NECESSARY ENTRIES")
        return "break"
    else:
        frame.tkraise()
        if frame == starting_FRAME:
            if var7.get()!=1:
                #root.geometry("275x230+650+0")
                if var9.get()==1 or  var10.get()==1:
                    root.geometry("435x323+650+0")
                else:
                    root.geometry("287x323+650+0")
                    
            else:
                #root.geometry("275x340+650+0")
                if var9.get()==1 or  var10.get()==1:
                    root.geometry("435x427+650+0")
                else:
                    root.geometry("287x427+650+0")
                    
               
        if frame == PIPE_FRAME_FRAME:
            if var8.get()!=1:
                root.geometry("470x105+550+0")            
            else:
                root.geometry("470x235+550+0")
                
        frame.grid(row =0,column=0,sticky='nsew')

        
def check6(event):
    global j
    if var8.get()!=1:
        #j = int(j_variable.get())
        observation_entrey.grid(row = 8 , column= 0,padx = 5, pady = 5, sticky="W",columnspan = 10)
        devlabel.grid(row = 10 , column= 0, columnspan = 10, pady = 5, padx= 0)
        root.geometry("470x235")
        print("var7=", var7.get(),"j=",j)
        
                
    else:
        observation_entrey.grid_forget()
        devlabel.grid_forget()
        root.geometry("470x105")
        print("var7=", var7.get(),"j=",j)

        

def j_variablefunc():
    global j,insertion_var,k
    if var7.get()==1:
        j = int(jr_variable.get())
        k= int(kr_variable.get())
        j_variable_btn.config(text="L_{} R_{}".format(j,k),bg="green2")
        insertion_var =False
        print("var7=","k=",k, var7.get(),"j=",j)



def quitt():
    root.quit()
    sys.exit()
    #top = Toplevel()


def load_list(list_of_defcts):
    try:
        text_default=open(list_of_defcts,'r')
        content=(text_default.read()).split("\n")
        text_default.close()
        print("content=",content)
    except:
        text_default=open(list_of_defcts,'w')
        if list_of_defcts=="defaut_AREP.txt":
            text_default.write("AA\nBA\nBU\nDL\nF\nSCVE\n")
        if list_of_defcts=="defaut_AMEULER.txt":
            text_default.write("AA\nBA\nBU\nDL\nF\nSCVE\n")
        if list_of_defcts=="defaut_ACHUTE.txt":
            text_default.write("AA\nBA\nBU\nDL\nF\nSCVE\n")
        if list_of_defcts=="defaut_OK.txt":
            text_default.write("AA\nBA\nBU\nDL\nF\nSCVE\n")
        if list_of_defcts=="PATH_CLIENT.txt":
            text_default.write(r"C:\Users\111\Desktop\client")
            
        if list_of_defcts=="LIST_INSPECTERUS.txt":
            text_default.write('BOUZID YASSINE\nBOUZID YASSINE\nBOUZID YASSINE\n')
            
        if list_of_defcts=="PATH_RAPPORT.txt":
            text_default.write(r"C:\RAPPORTS")
            
        if list_of_defcts=="LISTE_Projets.txt":
            text_default.write('CEEG KD/AL\nR-GZ2')
        
            
                            
        text_default.close()
        text_default=open(list_of_defcts,'r')
        content=(text_default.read()).split("\n")
        text_default.close()
    return content

def covert_to_excel(listing):
    listing =  (str(listing)).replace('[','')
    listing =  listing.replace(']','')
    listing =  listing.replace("'","")
    listing =  listing.replace("\n","")
    listing =  listing.replace(",","")
    return listing



def open_and_create_folder(event):
    
    global getname,NAME_REPORT
    
    pipe_name=str(PIPE_NAME.get())
    if PIPE_NAME.get().upper()=="TEST":
        print("test is test")
        getname='TEST'

    elif PIPE_NAME.get()=="" or pathE.get()=="":
        print("File up pipe name, path entries and defect name ")
        devlabel.config(text= "File up pipe name,\n path entries and defect name!!",fg="red",bg="yellow")
        return
    elif len(pipe_name)==2:
        print("length====",len(pipe_name),pipe_name[1])
        #A0001
        getname = (pipe_name[0]+"0"+"0"+"0"+pipe_name[1]).upper()
        print("getname",getname)
    elif len(PIPE_NAME.get())==3:
        #A0012
        getname = (pipe_name[0]+"0"+"0"+pipe_name[1]+pipe_name[2]).upper()
    elif len(PIPE_NAME.get())==4:
        #A0123
        getname = (pipe_name[0]+"0"+pipe_name[1]+pipe_name[2]+pipe_name[3]).upper()
    elif len(PIPE_NAME.get())==5:
        #A1234
        devlabel.config(text= "DEVELOPED BY BOUZID YASSINE \n CND-INSPECTOR RT-II 2021",fg="black",bg=color)
        getname = str(PIPE_NAME.get()).upper()
        print("getname= ",getname)
    elif ((PIPE_NAME.get()).upper().find('-BIS') != -1):
        devlabel.config(text= "BIS TUBE!!!",fg="black",bg="yellow")
        getname = str(PIPE_NAME.get()).upper()
    
    if not os.path.exists(pathE.get()):
        os.makedirs(pathE.get())
        devlabel.config(text= "REPORT FOLDER SUCCESSFULLY CREATED!!!",fg="black",bg="yellow")

    NAME_REPORT="{}_RAPPORT_UT_N_{}_{}_{}_AND_{}_{}".format(PROJECT,k,POST1,operators_names1,operators2_names1,time_value)
    #NAME_REPORT=str(NAME_REPORT).encode('ascii', 'ignore').decode('ascii')
    NAME_REPORT=NAME_REPORT.replace(' ','_')
    NAME_REPORT=NAME_REPORT.replace('-','_')
    NAME_REPORT=NAME_REPORT.replace('"','')
    print("name report=",NAME_REPORT)
    
    os.startfile(pathE.get())
    PIPE_NAME.delete(0,"end")
    PIPE_NAME.insert(0,getname)
    ROW = select_by_pipe_name(NAME_DATABASE,NAME_TABLE,NAME_REPORT,getname)
    #print(ROW)
    if ROW:
        rowid=ROW[0][0]
        #print("row id =====",rowid)
        print("messagebox this file is exists")
        messagebox.showinfo("FOLDER EXISTES!","THIS FOLDER IS EXISTES {}".format([R[0] for R in ROW]))
        






root = Tk()
root.config(bg ="white")
root.focus_force()
root.rowconfigure(0,weight= 1)
root.columnconfigure(0,weight=1)

#color = "light sky blue"
color = "spring green"
btncolor ="gold"
font_button=("Helvetica",10,"bold")
font_label=("Helvetica",9,"bold")
font_frame=("Helvetica",10,"bold")
font_entries=("Helvetica",15)

starting_FRAME = tk.Frame(root, width=100, height=100, background=color)
starting_FRAME.grid(row =0,column=0,sticky='nsew')

starting = LabelFrame(starting_FRAME, text = "INFO",  width = 35,height =20,font =font_frame,bg =color )
starting.grid(row =0,column=0,padx=2)


PIPE_FRAME_FRAME = tk.Frame(root, width=50, height=50, background=color )

PIPE_ET_KRACHAGE = LabelFrame(PIPE_FRAME_FRAME, text = "CONTROLE UT:",  width = 35,height =20,font =font_frame ,bg =color,labelanchor = "n")
PIPE_ET_KRACHAGE.grid(row =0,column=0,sticky='nsew', padx=5,pady = 5)
   
################################################################################################ STARTING FRAME ##########################################################################################

#OPERATOR1 LABEL

operator1_names_lab = Label(starting, text= "UT-INSP 01:",font =font_label, bg =color)
operator1_names_lab.grid(row = 0 , column= 0, padx = 10,sticky="W", pady=10)



# combobox5
def operators_namesfunc(event):
    global operators_names1
    operators_names1  = operators_names.get()
    operator1_names_lab.config(bg= "green2")
    print(operators_names1)

def operators_namesdel(event):
    global operators_names1
    operators_names1  = ""
    print("operators_names1 is deleted!")

## Adding combobox FILM A REFAIR
n4 = tk.StringVar() 
 
operators_names = ttk.Combobox(starting, width = 15, textvariable = n4,font= font_label ,stat="readonly")

operators_names['values'] =load_list("LIST_INSPECTERUS.txt")
   
  
operators_names.grid(row = 0,column = 1, padx = 10, pady=10) 
operators_names.current()
operators_names.bind("<<ComboboxSelected>>", operators_namesfunc)
operators_names.bind("<BackSpace>", operators_namesdel)
operators_names.focus_set()

# combobox6


#OPERATOR2 LABEL

operator2_names_lab = Label(starting, text= "UT-INSP 02:",font =font_label, bg =color)
operator2_names_lab.grid(row = 2 , column= 0, padx = 10,sticky="W", pady=10)


# combobox5
def operators2_namesfunc(event):
    global operators2_names1
    operators2_names1  = operators2_names.get()
    operator2_names_lab.config(bg= "green2")
    print(operators2_names1)

def operators2_namesdel(event):
    global operators2_names1
    operators2_names1  = ""
    print("operators2_names1 is deleted!")

## Adding combobox FILM A REFAIR
n4 = tk.StringVar() 
 
operators2_names = ttk.Combobox(starting, width = 15, textvariable = n4,font= font_label ,stat="readonly")

operators2_names['values'] =load_list("LIST_INSPECTERUS.txt")
  
operators2_names.grid(row = 2,column = 1, padx = 10, pady=10) 
operators2_names.current()
operators2_names.bind("<<ComboboxSelected>>", operators2_namesfunc)
operators2_names.bind("<BackSpace>", operators2_namesdel)

##################################################################################  PASSWORD   ################################################################################################################################################################################################################

########################################################################### PASSWORD DATABASE TABLE #############################################################################################

def insert_password(REC1,REC2):
    cnn =sq.connect('{}.db'.format(NAME_DATABASE))
    c = cnn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS PASSWORD_TABLE (INSPECTOR text, PASSWORD text)")
    cnn.commit()
    cnn.close()
    print("table created done!")
    cnn =sq.connect('{}.db'.format(NAME_DATABASE))
    c = cnn.cursor()
    c.execute("SELECT PASSWORD FROM PASSWORD_TABLE WHERE INSPECTOR = '{}'".format(REC1))
    rows = c.fetchall()
    if rows:
        print("message inspecteru deja exist")
    else:
        c.execute("INSERT INTO PASSWORD_TABLE VALUES (?,?)",(REC1,REC2))
        cnn.commit()
        cnn.close()
    print("done!")
    

def get_password(CURRENT):
    
    cnn =sq.connect('{}.db'.format(NAME_DATABASE))
    c = cnn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS PASSWORD_TABLE (INSPECTOR text, PASSWORD text)")
    cnn.commit()
    cnn.close()
    print("table created done!")
    
    cnn =sq.connect('{}.db'.format(NAME_DATABASE))
    c = cnn.cursor()
    c.execute("SELECT PASSWORD FROM PASSWORD_TABLE WHERE INSPECTOR = '{}'".format(CURRENT))
    rows = c.fetchall()
    print("rows========",len(rows))
    for frow in rows:
        row = frow
        print(row[0])
        print("")
    return rows

def update_password(insp,password):
    cnn =sq.connect('{}.db'.format(NAME_DATABASE))
    c = cnn.cursor()
    c.execute("SELECT rowid FROM PASSWORD_TABLE WHERE INSPECTOR = '{}'".format(insp))
    rows = c.fetchall()
    print("rows in update function is =",rows)
    if rows!=[]:
        c.execute("UPDATE PASSWORD_TABLE SET  INSPECTOR = (?), PASSWORD = (?) where rowid = (?)",(insp,password,rows[0][0]))
        print("done! update_password")
    cnn.commit()
    cnn.close()
    print("done!")
    
def update_password_function():
    
    if var9.get()==1:
        print("operators_names1 == ",operators_names1)
        password_of_first_inspector = get_password(operators_names1)
        #
        if password_of_first_inspector == []:
            print("password dosn't exist insert one !!!")
            insert_password(operators_names1,str(current_pass.get()))
            
            current_pass.delete(0,"end")
            new_pass1.delete(0,"end")
            new_pass2.delete(0,"end")
            
            messagebox.showinfo("PASSWORD INSERTED","PASSWORD INSERTED SUCCESSFULLY")
            
        elif (current_pass.get()== password_of_first_inspector [0][0]):
            print("current password =",current_pass.get(),"get_password(operators_names1)==",password_of_first_inspector[0][0])
            if( new_pass1.get()==new_pass2.get()):
                update_password(operators_names1,new_pass1.get())
                
                current_pass.delete(0,"end")
                new_pass1.delete(0,"end")
                new_pass2.delete(0,"end")
                
                messagebox.showinfo("PASSWORD UPDATED","PASSWORD UPDATED SUCCESSFULLY")
            else:
                print("new passwords do not match")
                messagebox.showinfo("PASSWORD ISSUE!","new passwords do not match")
       
        else:
            
            print("password for first insector = ",password_of_first_inspector[0][0],"current_pass =",current_pass.get())
            print("incorrect password")
            messagebox.showinfo("PASSWORD ISSUE!","incorrect password")

            
    if var10.get()==1:
        password_of_second_inspector = get_password(operators2_names1)
        print("password for second insector = ",password_of_second_inspector)
        
        if password_of_second_inspector == []:
            print("password dosn't exist insert one !!!")
            insert_password(operators2_names1,str(current_pass.get()))
            
            current_pass.delete(0,"end")
            new_pass1.delete(0,"end")
            new_pass2.delete(0,"end")
            
            messagebox.showinfo("PASSWORD INSERTED","PASSWORD INSERTED SUCCESSFULLY")
            
        elif (current_pass.get()==password_of_second_inspector[0][0]):
            print("current password =",current_pass.get(),"get_password(operators_names1)==",password_of_second_inspector[0][0])
            if( new_pass1.get()==new_pass2.get()):
                update_password(operators2_names1,new_pass1.get())
                
                current_pass.delete(0,"end")
                new_pass1.delete(0,"end")
                new_pass2.delete(0,"end")
                
                messagebox.showinfo("PASSWORD UPDATED","PASSWORD UPDATED SUCCESSFULLY")
            else:
                print("new passwords do not match")
                
                messagebox.showinfo("PASSWORD ISSUE!","new passwords do not match")
                
        
            
        else:
            print("password for second insector = ",password_of_second_inspector[0][0]," and current_pass =",current_pass.get())
            print("incorrect passowrd")
            messagebox.showinfo("PASSWORD ISSUE!","incorrect password")

    current_pass.delete(0,"end")
    new_pass1.delete(0,"end")
    new_pass2.delete(0,"end")
            
########################################################################################################################################################################


##def check_insp_FUNC_1(event):
##    if var9.get()!=1:
##        
##        check_up_pass1.grid(row = 0 , column= 0,padx = 5, pady = 10,sticky="W")
##    else:
##        
##        check_up_pass1.grid_forget()
##        
##
##
##def check_insp_FUNC_2(event):
##    if var10.get()!=1:
##        
##        check_up_pass2.grid(row = 3 , column= 0,padx = 5, pady = 10,sticky="W")
##    else:
##       
##       check_up_pass2.grid_forget()

       
def show_update_password(event):
    if var9.get()!=1 or  var10.get()!=1:
        
        current_pass_lab.grid(row = 0 , column= 5, padx = 10,sticky="W", pady=10)
        current_pass.grid(row = 1    , column= 5,padx = 5, pady = 10,sticky="W")
        
        new_pass1_lab.grid(row = 2 , column= 5, padx = 10,sticky="W", pady=10)
        new_pass1.grid   (row = 3    , column= 5,padx = 5, pady = 10,sticky="W")
        
        new_pass2_lab.grid(row = 4 , column= 5, padx = 10,sticky="W", pady=10)
        new_pass2.grid   (row = 5   , column= 5,padx = 5, pady = 10,sticky="W")

        update_password_btn.grid(row =9,column=5, padx = 10,pady = 5,sticky="E")
        
        if var7.get()==1:
            root.geometry("435x427")
            
        if var7.get()!=1:
            root.geometry("435x323")
            
        
    if var9.get()==1 or  var10.get()==1:
        current_pass_lab.grid_forget()
        current_pass.grid_forget()
        new_pass1_lab.grid_forget()
        new_pass1.grid_forget()
        new_pass2_lab.grid_forget()
        new_pass2.grid_forget()
        update_password_btn.grid_forget()
        current_pass.delete(0,"end")
        new_pass1.delete(0,"end")
        new_pass2.delete(0,"end")
        
        if var7.get()==1:
            root.geometry("287x427")
            
        if var7.get()!=1:
            root.geometry("287x323")



       
def GET_CURRENT_INSPECTEUR1(event):
    global CURRENT_INSPECTEUR
    if get_password(operators_names1)==[]:
        messagebox.showinfo("PASSWORD ISSUE","PASSWORD dosen't exist!\n please insert a new one")
        return "break"
        
    if PASS_INP_01.get()==get_password(operators_names1)[0][0]:
        CURRENT_INSPECTEUR = operators_names1
        print("current inspector is ***********",CURRENT_INSPECTEUR)
        
        check_update_passord1.config(bg="gold")#text = "{}".format(CURRENT_INSPECTEUR),
        check_update_passord2.config(bg=color)
        PASS_INP_02.delete(0,"end")
        #PASS_INP_01.grid_forget()
        #check_up_pass1.grid_forget()
    else:
        check_update_passord1.config(bg="red")
        CURRENT_INSPECTEUR=None
        messagebox.showinfo("PASSWORD ISSUE","INCORRECT PASSWORD!")
        
def GET_CURRENT_INSPECTEUR2(event):
    global CURRENT_INSPECTEUR
    if get_password(operators2_names1)==[]:
        messagebox.showinfo("PASSWORD ISSUE","PASSWORD dosen't exist!\n please insert a new one")
        return "break"
    if PASS_INP_02.get()==get_password(operators2_names1)[0][0]:
        CURRENT_INSPECTEUR = operators2_names1
        print("current inspector is ---------------- ",CURRENT_INSPECTEUR)
        check_update_passord2.config(bg="gold")#text = "{}".format(CURRENT_INSPECTEUR),
        check_update_passord1.config(bg=color)
        PASS_INP_01.delete(0,"end")
        #PASS_INP_02.grid_forget()
        #check_up_pass2.grid_forget()
    else:
        check_update_passord2.config(bg="red")
        CURRENT_INSPECTEUR=None
        messagebox.showinfo("PASSWORD ISSUE","INCORRECT PASSWORD!")
        


############ passowrd for inspectORS:
var7 = IntVar()
var9 = IntVar()
var10 = IntVar()
#var11 = IntVar()
#var12 = IntVar()
var13 = IntVar()


PASS_INP_01 = Entry(starting, width = 18,relief ="sunken",font= font_label, bg ="white",show="*")
PASS_INP_01.grid(row = 1 , column= 1,padx = 10, pady = 10,sticky="W")
PASS_INP_01.bind("<Return>", GET_CURRENT_INSPECTEUR1)


PASS_INP_02 = Entry(starting, width = 18,relief ="sunken",font= font_label, bg ="white",show="*")
PASS_INP_02.grid(row = 3 , column= 1,padx = 10, pady = 10,sticky="W")
PASS_INP_02.bind("<Return>", GET_CURRENT_INSPECTEUR2)
#check_up_pass1= Checkbutton(starting, text = "update password 1",font =font_label, variable = var11, bg =color)
#check_up_pass1.bind('<Button-1>',show_update_password)

########################################### update password ################################################

check_update_passord1= Checkbutton(starting, text = "UPDATE PASS 1",font =font_label, variable = var9, bg =color)
check_update_passord1.grid(row = 1 , column= 0,padx = 5, pady = 10,sticky="W")
check_update_passord1.bind('<Button-1>',show_update_password)

check_update_passord2= Checkbutton(starting, text = "UPDATE PASS 2",font= font_label, variable = var10, bg =color)
check_update_passord2.grid(row = 3 , column= 0,padx = 5, pady = 10,sticky="W")
check_update_passord2.bind('<Button-1>',show_update_password)


#check_up_pass2= Checkbutton(starting, text = "update password 2",font =font_label, variable = var12, bg =color)
#check_up_pass2.bind('<Button-1>',show_update_password)


current_pass_lab = Label(starting, text= "OLD PASSWORD",font =font_label, bg ="yellow")
current_pass = Entry(starting, width = 18,relief ="sunken",font= font_label, bg ="white",show="")

new_pass1_lab = Label(starting, text= "NEW PASSWORD:",font =font_label, bg ="yellow")
new_pass1 =Entry(starting, width = 18,relief ="sunken",font= font_label, bg ="white",show="")

new_pass2_lab = Label(starting, text= "REPEAT PASSWORD:",font =font_label, bg ="yellow")
new_pass2 = Entry(starting, width = 18,relief ="sunken",font= font_label, bg ="white",show="")

update_password_btn= tk.Button(starting,text="UPDATE",bg = "yellow",font =font_button ,width = 10,command=update_password_function)


################################################################################################################################################################################################################################################################################################################################################################################################################################################









# POST LABEL

POST_lab = Label(starting, text= "POSTE:",font =font_label, bg =color)
POST_lab.grid(row = 4 , column= 0, padx = 10,sticky="W", pady=10)


def POSTfunc(event):
    global POST1
    POST1  = POST.get()
    POST_lab.config(bg = "green2")
    
    print(POST1)

def POSTdel(event):
    global POST1
    POST1  = ""
    print("POST1 is deleted!")

n5 = tk.StringVar() 
 
POST = ttk.Combobox(starting, width = 15, textvariable = n5,font= font_label ,stat="readonly")

POST['values'] =('A  1 er',  
                 'A  2 eme', 
                 'A  3 eme',
                 'B  1 er',  
                 'B  2 eme', 
                 'B  3 eme',
                 'C  1 er',  
                 'C  2 eme', 
                 'C  3 eme',
                 'D  1 er',  
                 'D  2 eme', 
                 'D  3 eme')
  
POST.grid( row = 4,column = 1,padx = 10, pady=10) 
POST.current()
POST.bind("<<ComboboxSelected>>", POSTfunc)
POST.bind("<BackSpace>", POSTdel)


# EQUIPE

# EQUIPE LABEL

PROJECT_lab = Label(starting, text= "Projet:",font =font_label, bg =color)
PROJECT_lab.grid(row = 5 , column= 0, padx = 10,sticky="W", pady=10)


def PROJECTfunc(event):
    global PROJECT
    if str(EQUIPE.get())!="":
        PROJECT  = str(EQUIPE.get()).replace('/','_')
        PROJECT  = PROJECT.replace('-','_')
        PROJECT  = PROJECT.replace(' ','_')
        
        PROJECT_lab.config(bg = "green2")
    else:
        PROJECT_lab.config(bg = "orange1")
        
    print(PROJECT)

def PROJECTdel(event):
    global PROJECT
    PROJECT  = ""
    print("PROJECT is deleted!")

n5 = tk.StringVar() 
 
EQUIPE = ttk.Combobox(starting, width = 15, textvariable = n5,font= font_label ,stat="readonly")

EQUIPE['values'] =load_list("LISTE_Projets.txt")

EQUIPE.grid(row = 5,column = 1, padx = 10, pady=5,sticky="W") 
EQUIPE.current()
EQUIPE.bind("<<ComboboxSelected>>", PROJECTfunc)
EQUIPE.bind("<BackSpace>", PROJECTdel)




pathE = Entry(starting, width = 19,relief ="groove", font =font_entries ,bg="white")
#pathE.grid(row = 4 , column= 1, pady = 10,columnspan = 3)
#pathE.insert(0,"\\\poste2-rx\\production2\\CEEG 2020")
pathE.insert(0,r"{}".format(covert_to_excel(load_list("PATH_RAPPORT.txt"))))

pathE_lab = Label(starting, text= "PATH:",font =font_label, bg =color)
#pathE_lab.grid(row = 4 , column= 0, padx = 10,sticky="W")

var7 = IntVar()

continueing_checkbtn= Checkbutton(starting, text = "R-INCOMPLET",font =font_label, variable = var7, bg =color)
continueing_checkbtn.grid(row = 9 , column= 0,padx = 5, pady = 10,sticky="W")
continueing_checkbtn.bind('<Button-1>',check5)

fr1_btn= tk.Button(starting,text="ENTER",bg = "yellow",font =font_button ,width = 10,command=lambda:show_frame(PIPE_FRAME_FRAME))
fr1_btn.grid(row =9,column=1, padx = 10,pady = 5,sticky="E")

#j_variable_lab = Label(starting, text= "line N°:",font =("Helvetica",10,"bold"), bg =color)
#j_variable = Spinbox(starting,from_=1, to = max_line ,bg ="white",increment =1,width = 6, font =("Helvetica",15),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")
#j_variable.grid(row = 4 , column= 1, pady = 10,columnspan = 2)
#j_variable.delete(0,"end")
#j_variable.insert(0,"8")
#j_variable.bind("<<SpinboxSelected>>",j_variablefunc)

j_variable_btn= tk.Button(starting,text="VALIDE",width = 10,font =font_button,bg = "GOLD",command=j_variablefunc)

jr_variable_lab = Label(starting, text= "Line N°:",font =font_label, bg =color)
kr_variable_lab = Label(starting, text= "Rapport N°:",font =font_label, bg =color)
jr_variable = Spinbox(starting,from_=1, to = (max_line-1) ,bg ="white",increment =1,width = 3, font =font_label,buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")
kr_variable = Spinbox(starting,from_=1, to = 10 ,bg ="white",increment =1,width = 3, font =font_label,buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")



################################################################################################ END STARTING FRAME ##########################################################################################


PIPE_NAME_lab = Label(PIPE_ET_KRACHAGE, text= "PIPE: ",font =font_label, bg =color)
PIPE_NAME_lab.grid(row = 0 , column= 0, padx = 5,pady = 5)

# pipe name 
PIPE_NAME = Entry(PIPE_ET_KRACHAGE, width = 6,relief ="sunken", font =font_entries, bg ="white")
PIPE_NAME.grid(row = 0 , column= 1,padx = 5, pady = 5)
PIPE_NAME.focus_set()
PIPE_NAME.bind("<Return>", open_and_create_folder)

# krachage
krash_lab = Label(PIPE_ET_KRACHAGE, text= "SNP_K:",font =font_label, bg =color)
krash_lab.grid(row = 0 , column=2, padx = 5)

krash = Spinbox(PIPE_ET_KRACHAGE,from_=0, to = 20 ,bg ="white",increment =1,width = 2, font =("Helvetica",15),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")
krash.grid(row = 0, column= 3, padx = 5)
krash.delete(0,"end")
krash.insert(0,"0")


krash_OP_lab = Label(PIPE_ET_KRACHAGE, text= "OP_K:",font =font_label, bg =color)
krash_OP_lab.grid(row = 1 , column=2, padx = 5)

krash_OP = Spinbox(PIPE_ET_KRACHAGE,from_=0, to = 20 ,bg ="white",increment =1,width = 2, font =("Helvetica",15),buttonbackground = "orange" ,relief ="sunken", highlightcolor= "yellow")
krash_OP.grid(row = 1, column= 3, padx = 5)
krash_OP.delete(0,"end")


# buttons
#START_btn = Button(PIPE_ET_KRACHAGE, text = "START", bg =btncolor,activebackground="YELLOW1",font =font_button,height = 1, width = 10)
#START_btn.grid(row = 0 , column= 2,padx = 5, pady = 5,columnspan=1)


finish_tube = Button(PIPE_ET_KRACHAGE, text = "INSÈRE", bg =btncolor,activebackground="orange",font =font_button,height = 1, width = 10)
finish_tube.grid(row = 0 , column= 4,padx = 5, pady = 5)
finish_tube.bind('<Return>',tube_finished)
finish_tube.bind('<Button-1>',tube_finished)

report_closed_btn = Button(PIPE_ET_KRACHAGE, text = "CLÔTURE", bg =btncolor,activebackground="red",font =font_button,height = 1, width = 10)
report_closed_btn.grid(row = 0 , column= 5,padx = 5, pady = 5)
report_closed_btn.bind('<Return>',report_closed)
report_closed_btn.bind('<Button-1>',report_closed)

#QUIT_btn = Button(ACTIONS_FRAME, text = "QUIT", bg =btncolor,activebackground="RED",font =("Helvetica",10,"bold"),height = 2, width = 13, command = quitt)
#QUIT_btn.grid(row = 5 , column= 2,padx = 5, pady = 10)
#QUIT_btn.bind('<Return>',quitt)
#QUIT_btn.bind('<Button-1>',quitt)
# line 3
devlabel = Label(PIPE_ET_KRACHAGE, text= "DEVELOPED BY BOUZID YASSINE \n CND-RT-II 2021",font =("Algerian",10,"bold"), bg =color)
var8 = IntVar()

observation_checkbtn= Checkbutton(PIPE_ET_KRACHAGE, text = "OBS:",font =font_label, variable = var8, bg =color)
observation_checkbtn.grid(row = 1 , column= 0,padx = 5, pady = 3,sticky="W")
observation_checkbtn.bind('<Button-1>',check6)

QR_confirmed=False
def check_QR(event):
    global QR_confirmed
    if var13.get()==0:
        print_QR_checkbtn.config(bg ="orange")
        QR_confirmed=True
        
    elif var13.get()==1:
        
        print_QR_checkbtn.config(bg=color)
        QR_confirmed=False
        
print_QR_checkbtn= Checkbutton(PIPE_ET_KRACHAGE, text = "QR",font =font_label, variable = var13, bg =color)
print_QR_checkbtn.grid(row = 1 , column= 1,padx = 5, pady = 3,sticky="W")
print_QR_checkbtn.bind('<Button-1>',check_QR)



observation_entrey  = ScrolledText(PIPE_ET_KRACHAGE, height = 3, width = 38,relief ="sunken", font =("Helvetica",15), bg ="white")


#j_variable.grid(row = 4 , column= 1, pady = 10,columnspan = 2)
#j_variable.delete(0,"end")
#j_variable.insert(0,"8")
#j_variable.bind("<<SpinboxSelected>>",j_variablefunc)

fr3_btn= tk.Button(PIPE_ET_KRACHAGE,text="RETOUR >>",bg = "yellow",width = 10,font =font_button,command=lambda:show_frame(starting_FRAME))
fr3_btn.grid(row = 1 , column= 5,padx = 5,pady = 5,sticky="E")



############################################################################################### END PIPE_FRAME ############################################################################################################




icon ="""AAABAAEAICAAAAEAIACoEAAAFgAAACgAAAAgAAAAQAAAAAEAIAAAAAAAABAAACMuAAAjLgAAAAAAAAAAAAAoFhX/OSko/zkpKP84KCf/QjMy/0Y3Nv88LCz/JhQT/ygWFf8nFhX/JxYV/ycWFf8nFhX/JxYV/ycWFf8nFhX/JxYV/ycWFf8nFhX/JxYV/ycWFf8nFhX/JxUU/zYlJf85KSn/Oiop/zcnJv84Jyb/Oiop/zwsK/85KSj/KhgX/xgFBP8uHRz/LBsa/y4dHP9BMjH/Oysq/zcnJv8XAwL/FgMC/xYDAv8WAwL/FgMC/xYDAv8WAwL/FgMC/xYDAv8WAwL/FgMC/xYDAv8WAwL/FgMC/xYDAv8WAwL/IhAP/yQSEf8lFBP/IhAP/yIQD/8nFRT/KhkY/ygXFv8ZBgX/rqen/6qjo/+rpKP/q6Sj/6ihoP+mnp3/qKCe/62lpP+tpaP/raWj/62lo/+tpaP/raWj/62lo/+tpaP/raWj/62lo/+tpaP/raWj/62lo/+tpaP/raWj/62lo/+qoqH/qaGg/6mhoP+qo6L/q6Sk/6uko/+qo6P/qqOj/66np/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////z7+//8/Pz/+/r5////////////3868/8utj//KrI7/yqyO/8qsjv/KrI7/yqyO/8qsjv/KrI7/yqyO/8qsjv/KrI7/yqyO/8qsjv/JqYv/yKmK/8mpi//KrI7/yqyO/8qsjv/LrY//3869////////////+/r5//z8/P/8+/v///////79/f//////8urj/59lLP+TUhH/pnA6/6ZwO/+mcDv/pnA7/6ZwO/+mcDv/pnA7/6ZwO/+mcDv/pnA7/6ZwO/+mcDv/pnA7/65+Tf+xglP/rXxL/6ZvOv+mcDv/pnE7/6ZwOv+TUhH/n2Ys//Lr5P///////v39/////////v7///////37+v+WVxj/sYFS////////////////////////////////////////////////////////////////////////////1Lqh/7qPZv/eyrf///////////////////////////+wgVL/l1cZ//38+/////////7+//38+///////zK6P/5dXGf////////////38+//+/fz//v38//79/P/+/fz//v38//79/P/+/fz//v38//79/P/9/Pv////////////GpIP/uI1j/4tEAP/HpoX///////z7+f/9/Pv///////////+WVxj/za6Q///////9/Pv//fz7//////+2iV3/soNU///////8+/n///////////////////////////////////////////////////7+///////18Or/yKeH/8elhP//////2cOs/51iJ//////////////////8+/n//////7GCVP+2il7///////38+//9/Pv//////7aKXv+xglP///////38+//////////////////////////////////////////////////+/fz///////n28/+yg1b/pnA6/45JBf/t4tj//fv6//79/P////////////38+///////sYFS/7aKX////////fz7//38+///////tope/7GCU////////fz7/////////////////////////////////////////////v7+///////v5dz/za+S/9zHs///////qXZC/7uSaf///////fz7/////////////fz7//////+xgVL/topf///////9/Pv//fz7//////+2il7/sYFS///////8+/n//v39//79/f/+/f3//v39//79/f///v7//////////v/9+/r//////9O5oP+fZCr/k1EP/6x6SP/7+Pb/9e/q///////+/fz//v39//79/f/8+/n//////7CBUv+2il////////38+//9/Pv//////7aJXf+zhVf////////////////////////////////////////////+/fz////////////////////////////9/Pv/j0sI/+fazP//////////////////////////////////////soRW/7aKXv///////fz7//38+///////t4xh/6p2Q//07uj/697S/+zh1f/s4dX/7OHW/+zh1v/r39P//Pr5///////17+r/49LB/+rd0P/u5Nr/69/T/+7j2P/bxrH/49PC//38+//z7Ob/697S/+zh1v/s4dX/697S//Tu6P+pdUL/t4xh///////9/Pv//fz7//////++l3D/hDkA/45JBP+MRgD/jEYA/4xGAP+MRgD/jEYA/4tEAP+YWhz/4c++///////QtJn/k1IR/4I1AP+LRAD/jUgD/5BNCv+JQQD/7uPZ/7mPZv+DOAD/jUgD/4xGAP+MRgD/jkkE/4Q5AP++l3D///////38+//9/Pv//////76XcP+EOQD/jkkE/4xGAP+MRgD/jEYA/4xGAP+MRgD/jEUA/4lBAP/u5Nv/z7KV/9vGsv/9/f3/yKiH/49LCP+EOQD/i0QA/4lBAP/u5Nr/uY9m/4M4AP+OSQP/jEYA/4xGAP+OSQT/hDkA/76XcP///////fz7//38+///////vpdw/4Q5AP+OSQX/jEYA/4xGAP+MRgH/jEYA/45KBf+CNgD/z7OX//Hp4v+EOQD/hDkA/6l1Qv/r4NX/9vHt/7uSaf+MRgD/fzIA/+7k2v+6kGb/hDgA/45JBP+MRgD/jEYA/45JBf+EOQD/vpdw///////9/Pv//fz7//////++l3D/hDkA/45JBP+MRgD/jEYA/4xGAP+NSAP/hjwA/6BmLf//////o2w1/4c+AP+PSwb/hz0A/4E1AP/CnXn///////Ps5f+0hln/7eLX/7mPZf+EOQD/jkkE/4xGAP+MRgD/jkkE/4Q5AP++l3D///////38+//9/Pv//////76XcP+EOQD/jkkE/4xGAP+MRgD/jEYB/4xGAP+LRAD/7uXb/8qpiv99LgD/kEwJ/41HAv+OSgX/hz0A/+XVxv/RtZn/sIFS//Pt5v//////t4tf/4Q5AP+OSQT/jEYA/4xGAP+OSQT/hDkA/76XcP///////fz7//38+///////vpdw/4Q5AP+OSQT/jEYA/4xGAP+NSAL/hz8A/6JpMf/8+vn/28ax/55kKv+DNwD/i0UA/4M4AP+9lW7//Pr5/45JBf+CNQD/ikMA/8OffP+sekn/hjwA/41IA/+MRgD/jEYA/45JBP+EOQD/vpdw///////9/Pv//fz7//////++l3D/hDkA/45JBP+MRgD/jEYA/4xGAP+OSQX/hDgA/7GCVP//////+vj2/9W9pP+TURD/kU4M//////+yhFX/hDkA/5BMCP+LRQD/hDoA/4pDAP+MRwH/jEYA/4xGAP+MRgD/jkkE/4Q5AP++l3D///////38+//9/Pv//////76XcP+EOQD/jkkE/4xGAP+MRgD/jEYB/4xGAP+JQgD/8+zl/9vGsf+QTgz/8uvk//7+/v/28e3/3sq3/4Q6AP+OSQP/jEYA/4xHAf+OSQT/jEcB/4xGAP+MRgD/jEYA/4xGAP+OSQT/hDkA/76XcP///////fz7//38+///////vpdw/4Q5AP+OSQT/jEYA/4xGAP+OSgX/gjcA/8akg//59vT/7eLY/9/Muf/w59//w598/9jCq/+cYCT/iEAA/41HAv+MRgD/jEYA/4xGAP+MRgD/jEYA/4xGAP+MRgD/jEYA/45JBP+EOQD/vpdw///////9/Pv//fz7//////++l3D/hDkA/45JBP+MRgD/jUgC/4g/AP+bXyL//////6dyPv+HPgD//fz7/97Ktv9/MQD/hjwA/4tEAP+NRwH/jEYA/4xGAP+MRgD/jEYA/4xGAP+MRgD/jEYA/4xGAP+MRgD/jkkE/4Q5AP++l3D///////38+//9/Pv//////76XcP+EOQD/jkkE/4xGAP+NSAP/hDoA/+fZy//axK7/eyoA/7CBUv/69/X/lVQU/4pDAP+OSgX/jEcB/4xGAP+MRgD/jEYA/4xGAP+MRgD/jEYA/4xGAP+MRgD/jEYA/4xGAP+OSQT/hDkA/76Xcf///////fz7//38+///////vZZu/4Q5AP+OSQT/jkkE/4M4AP+3i2D//f39/5BMCf+LRAD/+vj2/7+Ycv+DNwD/jkkF/4xGAP+MRgD/jEYA/4xGAP+MRgD/jEYA/4xGAP+MRgD/jEYA/4xGAP+MRgD/jEYA/45JBP+EOQD/vZZv///////9/Pv//fz6///////FooD/gzgA/49LB/+JQQD/j0sH//38+/+6kWf/dyUA/9K3nf/t4tf/hjwA/41HAv+MRgD/jEYA/4xGAP+MRgD/jEYA/4xGAP+MRgD/jEYA/4xGAP+MRgD/jEYA/4xGAP+MRgD/j0oG/4M4AP/Go4H///////38+v/+/f3//////+zh1v+KQwD/jEUA/4hAAP/gzrz/6t7T/4AzAP+sekj//////6FoMP+IQAD/j0oG/41IA/+NSAP/jUgD/41IA/+NSAP/jUgD/41IA/+NSAP/jUgD/41IA/+NSAP/jUgD/49KBv+KQwD/ikMA/+zi1////////v39///////+/fz//////9O5n/+GPQD/jEUA/9K3nP+YWRr/fC0A/8Cadf+6kWj/fjAA/4c/AP+GPAD/hjwA/4Y8AP+GPAD/hjwA/4Y8AP+GPAD/hjwA/4Y8AP+GPAD/hjwA/4Y8AP+GPAD/hDkA/4hAAP/TuZ////////79/P/////////////////+/f3//////+rd0f+6kWj/o2oz/6hzPv+seUf/pm85/6RsNv+reEX/qnZD/6p2Q/+qdkP/qnZD/6p2Q/+qdkP/qnZD/6p2Q/+qdkP/qnZD/6p2Q/+qd0P/qndD/6t3RP+7kmn/6t3R///////+/f3////////////////////////////+/fz//////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////v38///////////////////////////////////////+/fz//fz7//79/P/+/f3//v38//79/f/+/f3//v38//79/P/+/fz//v38//79/P/+/fz//v38//79/P/+/fz//v38//79/P/+/fz//v38//79/f/+/fz//fz7//79/P//////////////////////AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA="""


    
icondata= base64.b64decode(icon)
## The temp file is icon.ico
tempFile= "icon.ico"
iconfile= open(tempFile,"wb")
## Extract the icon
iconfile.write(icondata)
iconfile.close()
root.wm_iconbitmap(tempFile)
#top.wm_iconbitmap(tempFile)

## Delete the tempfile
os.remove(tempFile)

#top.title("CONFIGURATION")
root.title("NDT REPORTER")
root.geometry("287x323")
root.call('wm', 'attributes', '.', '-topmost', True)
#root.resizable(False,False)
root.protocol('WM_DELETE_WINDOW', quitt)
root.mainloop()

































